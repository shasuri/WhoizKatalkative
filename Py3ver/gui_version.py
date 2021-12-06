from log_parser import *
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter.ttk import *
from os.path import isfile
from openpyxl import Workbook


def gui_log_opener(path, encoding_type):
    log_file = open(path, 'r', encoding=encoding_type)
    return log_file


def gui_set_user(chat_room,user_name):
    #member_list = chat_room.memberList

    user_exist = chat_member_searcher(chat_room.memberList, user_name)

    user_info = chat_member_searcher(chat_room.memberList, '회원님')

    if not user_exist:
        pass
    else:
        user_info.name = user_exist.name
        user_info.invitedDate = user_exist.invitedDate

        chat_room.memberList.remove(user_exist)


def result_loader(file_path, encoding_type):
	logFile = gui_log_opener(file_path, encoding_type)

	logFilePresentLine = next(logFile)
	"""str : Load first line of log file with next().
	"""

	chatRoomInstance = chat_room_info_checker(logFilePresentLine)
	"""ChatMember : Set information of chatting room on 'ChatRoom' class, parsing first line.
	"""

	logFilePresentLine = next(logFile)

	chatRoomInstance.set_log_saved_date(date_checker(logFilePresentLine[9:]))
	"""Slice the useless part of second line and parsing the date part with 'date_checker' function, then set on instance.
	"""

	blank_line_passer(logFile, 2)
	"""Pass the blank lines.
	"""

	for logFilePresentLine in logFile:
	    line_type_checker(logFilePresentLine, chatRoomInstance)
	"""Parse every single chatting line using 'line_type_checker' function.
	"""

	logFile.close()
	"""Print information of chatting members.
	"""

	return chatRoomInstance


class App(Tk):
	def __init__(self):
		Tk.__init__(self)

		self.title("Katalkative GUI Beta")
		self.geometry("600x300+50+125")
		self.resizable(False,False)

		self.container = Frame(self)
		self.container.pack(expand=True,fill=BOTH)
		self.container.grid_rowconfigure(0, weight=1)#research!!!
		self.container.grid_columnconfigure(0, weight=1)

		self.intro = IntroPage(self.container,self)
		self.intro.grid(row=0,column=0,sticky="nsew")
		
		self.intro.tkraise()
 

class IntroPage(Frame):
	def __init__(self, super_frame, controller):
		Frame.__init__(self,super_frame)
		self.controller = controller

		descript_frame = Frame(self)#,bd=2,bg="red") 
		descript_frame.pack(fill=X,padx=10,pady=10)

		intro_label = Label(descript_frame,text="Katalkative GUI Beta. Select your kakaotalk log file from mobile device.")
		intro_label.pack(side=LEFT,fill=X)

	#-------------------------------------
		input_frame = Frame(self)#,bd=2,bg="green")
		input_frame.pack(fill=BOTH,padx=10,pady=20)

		path_label = Label(input_frame,text='File path : ')
		path_label.pack(side=LEFT)

		self.path_input = Entry(input_frame,text="default")
		self.path_input.pack(expand=True,fill=X,side=LEFT,padx=10)

		path_select = Button(input_frame,text="Select...",width=10,command=self.file_selector)
		path_select.pack(side=LEFT,padx=20)

	#-------------------------------------
		encoding_select_frame = Frame(self)
		encoding_select_frame.pack(fill=BOTH,padx=10,pady=50)

		encoding_label = Label(encoding_select_frame,text="File encoding : ")
		encoding_label.pack(side=LEFT)

		self.encoding_select = Combobox(encoding_select_frame, state="readonly")
		self.encoding_select['values'] = ('UTF-8','CP949','euc-kr','ANSI','unicode_escape','ASCII')
		self.encoding_select.current(0)
		self.encoding_select.pack(side=LEFT,padx=10)

		encoding_info_button = Button(encoding_select_frame,text="i",width=2,command=self.encoding_information)
		encoding_info_button.pack(side=LEFT)

	#-------------------------------------
		bottom_frame = Frame(self)#,bd=2,bg="blue")
		bottom_frame.pack(fill=X,side=BOTTOM,pady=10)

		next_button = Button(bottom_frame,text="Next",width=10,command=self.file_exist)
		next_button.pack(side=RIGHT,padx=30)

	#-------------------------------------
	def file_selector(self):
		file_path = filedialog.askopenfilename(initialdir = "./",title="Select log file",filetypes=(("Text files", "*.txt"),("All files", "*.*")))
		self.path_input.delete(0,END)
		self.path_input.insert(0,file_path)

	def file_exist(self):
		open_file = self.path_input.get()
		
		if(isfile(open_file)):
			messagebox.showinfo("File check","Opening "+self.path_input.get()+"...")

			try:
				self.next_step(open_file,self.encoding_select.get())
				self.next_page.tkraise()
			
			except UnicodeDecodeError as unicode_error:
				messagebox.showinfo("Error!","Error : "+str(unicode_error)+".\nTry with 'UTF-8' encoding.")

			except ValueError as value_error:
				if self.encoding_select.get() == "UTF-8":
					messagebox.showinfo("Error!","Error : "+str(value_error)+".\nCheck the log is right to form.")
				else:
					messagebox.showinfo("Error!","Error : "+str(value_error)+".\nCheck the log is right to form, or try with 'UTF-8' encoding.")

			except IndexError as index_error:
				messagebox.showinfo("Error!","Error : "+str(index_error)+".\nCheck the log is right to form.")

		else:
			messagebox.showinfo("File check","Wrong path...")

	def encoding_information(self):
		
		messagebox.showinfo("Encoding tip","Default encoding type of log file of kakaotalk is UTF-8.")

	def next_step(self, file_path, encoding_type):

		self.chat_room = result_loader(file_path, encoding_type)

		if self.chat_room.roomType == 'personal':
			self.next_page = ResultPage(self.controller.container, self.controller, self.chat_room)
		else:
			self.next_page = UserSelectPage(self.controller.container, self.controller, self.chat_room)
		self.next_page.grid(row=0,column=0,sticky="nsew")


class UserSelectPage(Frame):

	def __init__(self,super_frame,controller,chat_room):
		Frame.__init__(self,super_frame)
		self.controller = controller
		self.chat_room = chat_room
		
		select_frame = Frame(self)
		select_frame.pack(fill=BOTH,padx=10,pady=20,side=TOP)

		select_label = Label(select_frame,text="Select your name in the box.")
		select_label.pack(side=LEFT,padx=10)

		name_select_list = ()

		chat_room_members = self.chat_room.memberList

		for box_index in range(len(chat_room_members)):	
			if chat_room_members[box_index].name != '회원님':
				name_select_list += (chat_room_members[box_index].name,)
			else:
				pass

		self.user_name_select = Combobox(select_frame, state="readonly")
		self.user_name_select['values'] = name_select_list
		self.user_name_select.current(0)
		self.user_name_select.pack(side=LEFT,padx=10)

		#--------------------------------------------------
		button_frame = Frame(self)
		button_frame.pack(fill=BOTH,side=BOTTOM,pady=10)

		result_button = Button(button_frame,text="Next",width=10,command=self.result_step)
		result_button.pack(side=RIGHT,padx=30)

	def result_step(self):
		try:
			gui_set_user(self.chat_room,self.user_name_select.get())
		except:
			pass
		self.result = ResultPage(self.controller.container,self.controller,self.chat_room)
		self.result.grid(row=0,column=0,sticky="nsew")
		self.result.tkraise()


class ResultPage(Frame):

	def __init__(self,super_frame,controller,chat_room):
		Frame.__init__(self,super_frame)
		self.chat_room = chat_room
		self.controller = controller
		self.controller.geometry("600x375+100+100")
		self.controller.resizable(True,False)
		
		#------------------------------------------
		title_frame = Frame(self)
		title_frame.pack(fill=BOTH,padx=10,pady=10)
		
		if self.chat_room.roomType == 'personal':
			title_label = Label(title_frame,text="You are chatting with : "+self.chat_room.title)
		else:
			title_label = Label(title_frame,text="Chatting room title : "+self.chat_room.title)
		title_label.pack(side=LEFT)
		
		#------------------------------------------
		date_frame = Frame(self)
		date_frame.pack(fill=BOTH,padx=10)

		save_date_label = Label(date_frame,text="Log is saved at "+str(self.chat_room.logSaveDate))
		save_date_label.pack(side=LEFT)

		#-------------------------------------------
		bottom_frame = Frame(self)
		bottom_frame.pack(fill=BOTH,padx=10,pady=10)

		question_label = Label(bottom_frame,text="So, WhoizKatalkative?")
		question_label.pack(side=TOP)

		#-------------------------------------------
		info_frame = Frame(bottom_frame)
		info_frame.pack(expand=YES,fill=BOTH)

		self.info_value_table = []
		self.info_types = ('Invited date','Talk average','Talk','Plain texts','Chat length','Emoticons','Images','Internet links','Videos','Hashtags','Etc. files','Addresses','Voice records')

		self.result_table = Treeview(info_frame,columns=self.info_types,selectmode="extended")
		
		table_yscroll = Scrollbar(info_frame,orient="vertical",command=(self.result_table).yview)
		table_yscroll.pack(side=RIGHT,fill=Y)
		table_xscroll = Scrollbar(info_frame,orient="horizontal",command=(self.result_table).xview)
		table_xscroll.pack(side=BOTTOM,fill=X)
		
		self.result_table.configure(yscrollcommand=table_yscroll.set,xscrollcommand=table_xscroll.set)

		self.result_table.pack(expand=YES,fill=BOTH)

		self.result_table.column('#0',stretch=YES,minwidth=10,width=100)
		self.result_table.heading('#0',text='User name')
		
		for type_set_index in range(len(self.info_types)):
			self.result_table.column('#'+str(type_set_index+1),stretch=YES,minwidth=20,width=80)
			self.result_table.heading('#'+str(type_set_index+1),text=self.info_types[type_set_index],anchor=W)
		
		self.result_table.column('#1',stretch=YES,minwidth=20,width=130)

		# Consist table contents.
		for insert_index in range(len(self.chat_room.memberList)):
			info_values = (self.chat_room.memberList[insert_index].invitedDate,)

			#Calc average.
			if self.chat_room.memberList[insert_index].infoList[MemberInfo.talk.value] != 0 :
				talk_average = float(self.chat_room.memberList[insert_index].infoList[MemberInfo.talkSize.value]) / self.chat_room.memberList[insert_index].infoList[MemberInfo.talk.value]
				talk_average = round(talk_average,2)
			else :
				talk_average = 0.00

			info_values += (talk_average,)
			
			for value_index in range(len(self.chat_room.memberList[insert_index].infoList)):
				info_values += (self.chat_room.memberList[insert_index].infoList[value_index],)
			
			member_name = self.chat_room.memberList[insert_index].name
			self.result_table.insert('','end',text=member_name,values=info_values)
			
			info_values = (member_name,) + info_values			
			self.info_value_table.append(info_values)

		#-----------------------------------------
		
		end_frame = Frame(bottom_frame)
		end_frame.pack(fill=BOTH,side=BOTTOM)

		maxVal = 0
		for maxIndex in self.chat_room.memberList:
			if maxVal < maxIndex.infoList[MemberInfo.talkSize.value]:
				maxVal = maxIndex.infoList[MemberInfo.talkSize.value]
				maxWho = maxIndex.name

		who_label = Label(end_frame,text=maxWho+" is the most Katalkative!")
		who_label.pack(side=LEFT)

		exit_button = Button(end_frame,text="Exit",width=10,command=self.close_page)
		exit_button.pack(side=RIGHT,padx=10)

		extract_button = Button(end_frame,text="Extract",width=10,command=self.extract_select)
		extract_button.pack(side=RIGHT)
		
	def close_page(self):
		self.controller.destroy()

	def extract_select(self):
		self.extract_window = Toplevel(self.controller)
		(self.extract_window).geometry("300x120+50+125")
		(self.extract_window).resizable(False,False)
		(self.extract_window).focus_force()
		(self.extract_window).grab_set()

		extract_label = Label(self.extract_window,text="Choose which file type to extract.")
		extract_label.pack()

		self.var_sheet = BooleanVar()
		self.var_database = BooleanVar()

		checkbox_frame = Frame(self.extract_window)
		checkbox_frame.pack(fill=X,pady=10)

		check_sheet = Checkbutton(checkbox_frame,text="Spread sheet(.xlsx)",variable=self.var_sheet,onvalue=True,offvalue=False)
		check_sheet.pack(side=LEFT,padx=10)

		check_database = Checkbutton(checkbox_frame,text="Database(.sql)",variable=self.var_database,onvalue=True,offvalue=False)
		check_database.pack(side=LEFT,padx=10)

		choose_frame = Frame(self.extract_window)
		choose_frame.pack(side=BOTTOM,fill=X,pady=10)

		extract_cancel_button = Button(choose_frame,text="Cancel",width=10,command=self.extract_cancel)
		extract_cancel_button.pack(side=RIGHT,padx=10)

		extract_ok_button = Button(choose_frame,text="OK",width=10,command=self.extract_file)
		extract_ok_button.pack(side=RIGHT,padx=10)

	def extract_file(self):
		if not (self.var_sheet.get() or self.var_database.get()):
			messagebox.showinfo("Select","Select which file to extract on checkbox")

		else:
			if self.var_sheet.get():
				self.extract_sheet()
			
			if self.var_database.get():
				messagebox.showinfo("Extract as Database file","Sorry, not yet...(2018.07.26)")
			
			self.extract_cancel()

	def extract_cancel(self):
		self.extract_window.destroy()

	def extract_sheet(self):
		sheet_extract = Workbook()
		sheet_write = sheet_extract.active
		
		info_types_extract = ("User name",) + self.info_types

		for info_index in range(len(info_types_extract)):
			info_cell = sheet_write.cell(row=1,column=info_index+1)
			info_cell.value = info_types_extract[info_index]

		for member_index in range(len(self.info_value_table)):
			for info_value_index in range(len(self.info_value_table[member_index])):
				info_value_cell = sheet_write.cell(row=member_index+2,column=info_value_index+1)
				info_value_cell.value = self.info_value_table[member_index][info_value_index]

		extract_path = filedialog.asksaveasfilename(initialdir = "./",title = "Save as Spread sheet file",filetypes = (("Spread sheet files","*.xlsx"),("All files","*.*")),confirmoverwrite=False,defaultextension=".xlsx")

		if extract_path: #check if file path is null or not.
			sheet_extract.save(extract_path)
			
		sheet_extract.close


if __name__ == "__main__":
	Katalkative_beta = App()
	Katalkative_beta.mainloop()
